"""Splitter tests.

Same rule as test_workbook.py: run against the real workbook when it is present,
because the properties under test are properties OF that file — the two
dashboards whose row 1 names 1 column of 16,208, the `&` in a sheet name, the
`Lead Dump` header that is itself the rogue row.

The real sheets are split ONCE, in a module-scoped fixture, and every test that
needs output reads that. Splitting per test would mean writing `Lead Dump`'s
54,507 rows eight times over.
"""

from __future__ import annotations

import os
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from app.leads import normalize_identifier, row_to_lead
from app.split import (
    _UNNAMED,
    heads_a_table,
    lead_columns,
    named_columns,
    output_name,
    sheet_title,
    split_workbook,
)
from app.workbook import (
    build_columns,
    excel_serial_to_date,
    list_sheets,
    read_sheet,
    resolve_sheet,
    stream_leads,
)

WORKBOOK = Path(
    os.environ.get(
        "TEST_WORKBOOK",
        r"C:\Users\Diwakar.Adhikari01\Desktop\Businesses Dashboard Jun'26.xlsb",
    )
)

needs_workbook = pytest.mark.skipif(
    not WORKBOOK.exists(), reason=f"source workbook not present at {WORKBOOK}"
)


# ── filenames ──────────────────────────────────────────────────────────────


def test_output_names_are_safe_on_windows():
    # The real sheet names carry spaces and an ampersand; neither may reach a path.
    assert output_name(3, "BFL & BAU") == "03-bfl-bau.xlsx"
    assert output_name(6, "Login Data") == "06-login-data.xlsx"
    for name in ('a:b', 'a/b', 'a\\b', 'a?b', 'a*b', 'a"b', "a<b>", "a|b", "trailing. "):
        stem = output_name(1, name)[: -len(".xlsx")]
        # Nothing Windows bars in a filename, and no trailing dot or space —
        # which Windows silently strips, so two names could become one.
        assert set(stem) <= set("abcdefghijklmnopqrstuvwxyz0123456789-"), stem
        assert not stem.endswith(("-", ".", " ")), stem


def test_output_names_are_stable_and_cannot_collide():
    # Stable: a re-split must overwrite its own output, not accumulate copies.
    assert output_name(6, "Login Data") == output_name(6, "Login Data")
    # Two sheets that slug identically stay apart, because the ordinal is the
    # sheet's position and no two sheets share one.
    assert output_name(1, "Q1 Target") != output_name(2, "Q1  target")


def test_a_sheet_named_after_a_dos_device_is_still_openable():
    # CON, PRN, NUL, COM1 are reserved device names: Windows refuses to open a
    # file called CON.xlsx. The ordinal prefix is what stops one being produced.
    for device in ("CON", "PRN", "AUX", "NUL", "COM1", "LPT1"):
        stem = output_name(4, device).split(".")[0]
        assert stem.upper() not in {"CON", "PRN", "AUX", "NUL", "COM1", "LPT1"}


def test_a_sheet_name_that_sanitises_to_nothing_still_gets_a_name():
    assert output_name(2, "***") == "02-sheet.xlsx"


def test_sheet_title_stays_within_excels_rules():
    # Excel refuses these characters and caps a sheet name at 31 characters;
    # openpyxl raises rather than truncating, which would lose the whole sheet.
    assert sheet_title("Login Data") == "Login Data"
    assert "/" not in sheet_title("a/b:c*d?e[f]g")
    assert len(sheet_title("x" * 40)) == 31


# ── what counts as a table ─────────────────────────────────────────────────


def test_the_unnamed_marker_still_matches_build_columns():
    """`heads_a_table` counts build_columns' placeholder for a blank header.

    It matches the text rather than re-deriving it, so if that format ever
    changes this fails here — instead of every dashboard silently becoming a
    data sheet and being written out 16,000 columns wide.
    """
    assert _UNNAMED.match(build_columns(["A", None, "B"])[1])


def test_a_real_table_heads_itself():
    assert heads_a_table(["Apps_No", "FY", "Login_Date"])
    # Dash has a spacer column between its two side-by-side blocks: 17 of 18
    # named, and it is still a table.
    assert heads_a_table(["A"] * 17 + ["(column 18)"])


def test_a_dashboard_does_not():
    # Overall Dashboard's row 1: one title cell, then 16,379 blanks and a stray.
    assert not heads_a_table(["Product Type"] + [f"(column {i})" for i in range(2, 16_381)])
    assert not heads_a_table([])
    assert named_columns(["A", "(column 2)", "B"]) == 2


# ── against the real workbook ──────────────────────────────────────────────


@pytest.fixture(scope="module")
def split_once(tmp_path_factory):
    if not WORKBOOK.exists():
        pytest.skip(f"source workbook not present at {WORKBOOK}")
    out_dir = tmp_path_factory.mktemp("split")
    outcomes = split_workbook(WORKBOOK, out_dir)
    return out_dir, {o.sheet_name: o for o in outcomes}


@needs_workbook
def test_every_sheet_is_accounted_for(split_once):
    _, by_name = split_once
    assert len(by_name) == 10
    for outcome in by_name.values():
        # Written or refused, never absent and never silent.
        assert (outcome.path is None) != (outcome.reason is None)
        assert not outcome.failed, outcome.reason


@needs_workbook
def test_the_two_dashboards_are_refused_with_their_row_count(split_once):
    _, by_name = split_once
    for name in ("BFL & BAU", "Overall Dashboard"):
        outcome = by_name[name]
        assert outcome.path is None
        # The reason has to say how much was left behind. "Skipped" alone would
        # read as "there was nothing there", and there are 23 and 34 rows there.
        assert "row(s) left in the source" in outcome.reason
        assert "pivot or dashboard" in outcome.reason


@needs_workbook
def test_output_is_a_real_zip_container(split_once):
    out_dir, by_name = split_once
    for outcome in by_name.values():
        if outcome.path is None:
            continue
        # src/lib/import/actions.ts checks these four bytes before storing
        # anything, so an output that fails here cannot be uploaded at all.
        assert outcome.path.read_bytes()[:4] == b"PK\x03\x04"
        assert outcome.path.suffix == ".xlsx"
    # And nothing half-written was left behind under its temporary name.
    assert list(out_dir.glob("*.partial")) == []


@needs_workbook
def test_login_data_reads_back_identically(split_once):
    _, by_name = split_once
    source = read_sheet(WORKBOOK, "Login Data")
    produced = read_sheet(by_name["Login Data"].path, "Login Data")

    assert produced.columns == source.columns
    assert len(produced.columns) == 36
    assert produced.total_rows == source.total_rows == 1171
    assert produced.rows == source.rows


@needs_workbook
def test_identifiers_and_dates_survive_the_round_trip(split_once):
    _, by_name = split_once
    source = read_sheet(WORKBOOK, "Login Data", max_rows=1).rows[0]
    produced = read_sheet(by_name["Login Data"].path, "Login Data", max_rows=1).rows[0]

    # Apps_No is the identifier the whole portal joins on. It must not have been
    # written as text and come back in scientific notation.
    assert isinstance(produced["Apps_No"], float)
    assert normalize_identifier(produced["Apps_No"]) == normalize_identifier(source["Apps_No"])
    assert "e+" not in normalize_identifier(produced["Apps_No"])
    assert produced["SM_ID"] == source["SM_ID"] == "ICCSP90766"

    # A date stays a date, not a serial: calamine resolved it and openpyxl wrote
    # it back with a date number format.
    assert isinstance(produced["Login_Date"], date)
    assert produced["Login_Date"] == source["Login_Date"]


@needs_workbook
def test_a_resplit_file_is_still_recognised_on_re_upload(split_once):
    _, by_name = split_once
    # The sheet keeps its own name inside its own file, so resolve_sheet still
    # finds the transaction sheet without the admin picking it by hand.
    names = [s.name for s in list_sheets(by_name["Login Data"].path)]
    assert names == ["Login Data"]
    assert resolve_sheet(names, None) == "Login Data"


@needs_workbook
def test_duplicate_and_blank_headers_keep_the_names_they_were_given(split_once):
    _, by_name = split_once
    produced = read_sheet(by_name["Dash"].path, "Dash")
    # Dash carries two Location columns and a blank spacer. build_columns
    # disambiguated them on the way in and the file preserves that, so the two
    # do not collapse into one on the way back.
    assert "Location (1)" in produced.columns
    assert "Location (2)" in produced.columns
    assert "(column 9)" in produced.columns
    assert produced.columns == read_sheet(WORKBOOK, "Dash").columns


@needs_workbook
def test_computed_aggregates_lose_nothing_a_human_could_see(split_once):
    """openpyxl writes a float as %.16g, so the 17th digit does not survive.

    Asserted rather than waved away: it is a real, if tiny, loss, and a future
    writer that lost more than this should fail here rather than quietly
    rounding money. Excel itself carries 15 significant digits, so nothing
    anybody reads changes.
    """
    _, by_name = split_once
    source = read_sheet(WORKBOOK, "Dash")
    produced = read_sheet(by_name["Dash"].path, "Dash")

    for before, after in zip(source.rows, produced.rows):
        for key, value in before.items():
            if isinstance(value, float) and value:
                assert abs(after[key] - value) / abs(value) < 1e-15
            else:
                assert after[key] == value


# ── Lead Dump ──────────────────────────────────────────────────────────────


@needs_workbook
def test_lead_header_matches_what_stream_leads_keys_by():
    """The peeked header and stream_leads' keys must be the same list.

    `lead_columns` reads the header row itself because stream_leads does not
    hand its column names out, and the header has to be written before the first
    row is streamed. If the two ever disagree, `row.get(name)` starts missing and
    every value in the written file shifts a column — silently, because a missing
    key is just a blank cell.
    """
    columns = lead_columns(WORKBOOK, "Lead Dump")
    assert columns == list(next(iter(stream_leads(WORKBOOK))).keys())
    assert len(columns) == 15


@needs_workbook
def test_the_strays_never_reach_the_split_file():
    # Two stray labels sit at columns 16,381-2 reading "Product type" and
    # "Source". Narrowed before naming, the real columns keep their own names —
    # unnarrowed they come back as "Product type (1)" and "Source (1)", which is
    # what nulled both fields on all 54,507 leads.
    columns = lead_columns(WORKBOOK, "Lead Dump")
    assert "Product type" in columns and "Source" in columns
    assert "Product type (1)" not in columns and "Source (1)" not in columns
    assert "Location (1)" in columns and "Location (2)" in columns


@needs_workbook
def test_lead_dump_is_written_whole_and_at_its_real_width(split_once):
    _, by_name = split_once
    outcome = by_name["Lead Dump"]
    assert outcome.rows == 54_507
    assert outcome.columns == 15

    # Read back with openpyxl's read-only reader, one row at a time. read_sheet
    # cannot be used here: it refuses any sheet NAMED "Lead Dump", regardless of
    # how narrow the file actually is.
    workbook = openpyxl.load_workbook(outcome.path, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        assert workbook.sheetnames == ["Lead Dump"]

        rows = worksheet.iter_rows(values_only=True)
        header = list(next(rows))
        first = list(next(rows))
        assert header == lead_columns(WORKBOOK, "Lead Dump")
        assert len(first) == 15
    finally:
        workbook.close()

    source = next(iter(stream_leads(WORKBOOK)))
    assert first == [source.get(name) for name in header]

    # And the attribution the leads table is built from is unchanged.
    before, after = row_to_lead(source), row_to_lead(dict(zip(header, first)))
    assert after == before
    assert after.lead_no == "105683457"
    assert after.sm_code == "ICCS427343"


@needs_workbook
def test_a_lead_date_is_still_a_date_after_the_round_trip(split_once):
    """Register Date is written as the serial pyxlsb yielded, not as a date.

    Converting it here would mean deciding which floats are dates, and every NOP
    count and FRP amount between 1 and 60000 would be decided wrongly.
    `excel_serial_to_date` is the existing convention and it reads the serial out
    of the split file exactly as it reads it out of the source.
    """
    _, by_name = split_once
    workbook = openpyxl.load_workbook(by_name["Lead Dump"].path, read_only=True)
    try:
        rows = workbook[workbook.sheetnames[0]].iter_rows(values_only=True)
        header = list(next(rows))
        first = dict(zip(header, next(rows)))
    finally:
        workbook.close()

    assert excel_serial_to_date(first["Register Date"]) == date(2026, 6, 1)
    assert excel_serial_to_date(first["Register Date"]) == excel_serial_to_date(
        next(iter(stream_leads(WORKBOOK)))["Register Date"]
    )


# ── the writer's own hazards ───────────────────────────────────────────────


@needs_workbook
def test_a_value_that_looks_like_a_formula_stays_a_value(tmp_path):
    """openpyxl types any string starting with `=` as a formula.

    Written that way the cell carries `<f>` with no cached result, and every
    reader — calamine included — reads it as empty. The value is gone and
    nothing in the file says so, which is why _cell forces the string type.
    """
    from app.split import _write

    destination = tmp_path / "formulas.xlsx"
    _write(destination, "S", ["a", "b", "c"], [["=Term", "-", "+91 22"], ["@x", "", "plain"]])

    workbook = openpyxl.load_workbook(destination)
    try:
        assert [c.data_type for c in workbook["S"][2]] == ["s", "s", "s"]
        assert [c.value for c in workbook["S"][2]] == ["=Term", "-", "+91 22"]
    finally:
        workbook.close()

    assert read_sheet(destination, "S").rows[0]["a"] == "=Term"


def test_only_writes_the_sheets_it_was_asked_for(tmp_path):
    if not WORKBOOK.exists():
        pytest.skip(f"source workbook not present at {WORKBOOK}")

    outcomes = split_workbook(WORKBOOK, tmp_path, only=["manpower"])
    assert [o.sheet_name for o in outcomes] == ["Manpower"]
    # The ordinal is the sheet's position in the WORKBOOK, not in this run, so a
    # one-sheet re-split overwrites the file a full split wrote.
    assert outcomes[0].path.name == "08-manpower.xlsx"
    assert sorted(p.name for p in tmp_path.iterdir()) == ["08-manpower.xlsx"]
